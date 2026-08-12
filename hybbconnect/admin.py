from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin
from django.urls import path
from django.shortcuts import render, redirect
from django import forms

import pandas as pd

from .models import (
    CustomUser,
    Ticket,
    StaffPerformance,
    KitchenLog,
    Location,
    ClusterManagerProfile,
    SalarySlip,
    OrderPhoto,
)

from django.contrib.auth import get_user_model
from django.utils.html import format_html
import csv
import io
from django.shortcuts import render, redirect
from django.urls import path
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from .forms import UserBulkUploadForm
from .models import CustomUser
from django.contrib.auth.admin import UserAdmin
from openpyxl import load_workbook
from django.db import transaction

import csv
import pandas as pd
from django.contrib import admin, messages
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth import get_user_model
from django.utils.translation import gettext_lazy as _
from django.shortcuts import redirect, render
from django.urls import path
from django.contrib.auth.hashers import make_password

from .models import CustomUser, Location
from .forms import UserBulkUploadForm

User = get_user_model()


@admin.register(CustomUser)
class CustomUserAdmin(UserAdmin):

    change_list_template = "users_change_list.html"

    list_display = ("employee_id", "username", "role", "location",
                    "is_active", "is_staff")
    list_filter = ("role", "location", "is_active", "is_staff")
    search_fields = ("employee_id", "username", "email")
    ordering = ("employee_id",)

    fieldsets = UserAdmin.fieldsets + (
        ("Role & Location Details",
         {"fields": ("employee_id", "role", "location")}),
    )

    add_fieldsets = UserAdmin.add_fieldsets + (
        ("Role & Location Details",
         {"fields": ("employee_id", "role", "location")}),
    )

    # --------------------------
    # Custom URL
    # --------------------------
    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "bulk-upload/",
                self.admin_site.admin_view(self.bulk_upload_view),
                name="customuser_bulk_upload",
            ),
        ]
        return custom + urls

    # --------------------------
# Bulk Upload View
# Memory-Optimized Version
# --------------------------
    def bulk_upload_view(self, request):

        BATCH_SIZE = 500

        if request.method == "POST":

            form = UserBulkUploadForm(request.POST, request.FILES)

            if not form.is_valid():
                return render(
                    request,
                    "bulk_upload.html",
                    {"form": form}
                )

            file = request.FILES["file"]
            filename = file.name.lower()

            required = [
                "employee_id",
                "username",
                "role",
                "location",
                "password",
            ]

            created_count = 0
            skipped_count = 0

            # -------------------------------------------------
            # Load locations once
            # -------------------------------------------------
            locations = {
                str(location.code).strip(): location
                for location in Location.objects.all()
            }

            # -------------------------------------------------
            # Track duplicates inside the uploaded file
            # -------------------------------------------------
            uploaded_usernames = set()
            uploaded_employee_ids = set()

            # -------------------------------------------------
            # Process one batch at a time
            # -------------------------------------------------
            def process_batch(rows):

                nonlocal created_count, skipped_count

                if not rows:
                    return

                # ---------------------------------------------
                # Validate required columns
                # ---------------------------------------------
                for row_number, row in rows:

                    for field in required:

                        value = row.get(field)

                        if value is None or str(value).strip() == "":
                            messages.error(
                                request,
                                f"Row {row_number}: Missing or empty field: {field}"
                            )
                            raise ValueError("Invalid required field")

                # ---------------------------------------------
                # Get usernames / employee IDs in this batch
                # ---------------------------------------------
                usernames = set()
                employee_ids = set()

                for row_number, row in rows:

                    username = str(row["username"]).strip()
                    employee_id = str(row["employee_id"]).strip()

                    usernames.add(username)
                    employee_ids.add(employee_id)

                # ---------------------------------------------
                # Existing database records
                # ---------------------------------------------
                existing_usernames = set(
                    CustomUser.objects.filter(
                        username__in=usernames
                    ).values_list(
                        "username",
                        flat=True
                    )
                )

                existing_employee_ids = set(
                    CustomUser.objects.filter(
                        employee_id__in=employee_ids
                    ).values_list(
                        "employee_id",
                        flat=True
                    )
                )

                # ---------------------------------------------
                # Prepare users for bulk insert
                # ---------------------------------------------
                users_to_create = []

                for row_number, row in rows:

                    username = str(row["username"]).strip()
                    employee_id = str(row["employee_id"]).strip()
                    role = str(row["role"]).strip()
                    location_value = str(row["location"]).strip()
                    password = str(row["password"]).strip()

                    # -----------------------------------------
                    # Duplicate username
                    # -----------------------------------------
                    if (
                        username in existing_usernames
                        or username in uploaded_usernames
                    ):

                        skipped_count += 1

                        messages.warning(
                            request,
                            f"Row {row_number}: "
                            f"Skipped (username exists): {username}"
                        )

                        continue

                    # -----------------------------------------
                    # Duplicate employee ID
                    # -----------------------------------------
                    if (
                        employee_id in existing_employee_ids
                        or employee_id in uploaded_employee_ids
                    ):

                        skipped_count += 1

                        messages.warning(
                            request,
                            f"Row {row_number}: "
                            f"Skipped (employee ID exists): {employee_id}"
                        )

                        continue

                    # -----------------------------------------
                    # Check location
                    # -----------------------------------------
                    location_obj = locations.get(location_value)

                    if not location_obj:

                        messages.error(
                            request,
                            f"Row {row_number}: "
                            f"❌ Location not found: {location_value}"
                        )

                        raise ValueError(
                            f"Location not found: {location_value}"
                        )

                    # -----------------------------------------
                    # Mark as seen
                    # -----------------------------------------
                    uploaded_usernames.add(username)
                    uploaded_employee_ids.add(employee_id)

                    # -----------------------------------------
                    # Create user object
                    # -----------------------------------------
                    users_to_create.append(
                        CustomUser(
                            employee_id=employee_id,
                            username=username,
                            email=str(row.get("email") or "").strip(),
                            role=role,
                            location=location_obj,
                            password=make_password(password),
                            is_active=True,
                            is_staff=True,
                        )
                    )

                # ---------------------------------------------
                # Bulk insert
                # ---------------------------------------------
                if users_to_create:

                    with transaction.atomic():

                        CustomUser.objects.bulk_create(
                            users_to_create,
                            batch_size=BATCH_SIZE
                        )

                    created_count += len(users_to_create)

            # =================================================
            # CSV
            # =================================================
            if filename.endswith(".csv"):

                try:

                    text_file = io.TextIOWrapper(
                        file.file,
                        encoding="utf-8-sig",
                        newline=""
                    )

                    reader = csv.reader(text_file)

                    # -----------------------------------------
                    # Header
                    # -----------------------------------------
                    try:
                        headers = next(reader)
                    except StopIteration:
                        messages.error(
                            request,
                            "CSV file is empty."
                        )
                        return redirect("..")

                    headers = [
                        str(header)
                        .strip()
                        .lower()
                        .replace(" ", "_")
                        .replace("\ufeff", "")
                        for header in headers
                    ]

                    # -----------------------------------------
                    # Check required columns
                    # -----------------------------------------
                    missing_columns = [
                        field
                        for field in required
                        if field not in headers
                    ]

                    if missing_columns:

                        messages.error(
                            request,
                            "Missing required columns: "
                            + ", ".join(missing_columns)
                        )

                        return redirect("..")

                    batch = []

                    for row_number, values in enumerate(reader, start=2):

                        row = dict(
                            zip(headers, values)
                        )

                        batch.append(
                            (row_number, row)
                        )

                        if len(batch) >= BATCH_SIZE:

                            process_batch(batch)

                            batch.clear()

                    # -----------------------------------------
                    # Process remaining rows
                    # -----------------------------------------
                    if batch:
                        process_batch(batch)

                    text_file.detach()

                except UnicodeDecodeError:

                    messages.error(
                        request,
                        "CSV must be saved as UTF-8 encoding."
                    )

                    return redirect("..")

                except ValueError:

                    return redirect("..")

            # =================================================
            # Excel XLSX
            # =================================================
            elif filename.endswith(".xlsx"):

                try:

                    workbook = load_workbook(
                        file,
                        read_only=True,
                        data_only=True
                    )

                    worksheet = workbook.active

                    rows_iterator = worksheet.iter_rows(
                        values_only=True
                    )

                    # -----------------------------------------
                    # Header
                    # -----------------------------------------
                    try:
                        headers = next(rows_iterator)
                    except StopIteration:

                        messages.error(
                            request,
                            "Excel file is empty."
                        )

                        workbook.close()
                        return redirect("..")

                    headers = [
                        str(header or "")
                        .strip()
                        .lower()
                        .replace(" ", "_")
                        .replace("\ufeff", "")
                        for header in headers
                    ]

                    # -----------------------------------------
                    # Check required columns
                    # -----------------------------------------
                    missing_columns = [
                        field
                        for field in required
                        if field not in headers
                    ]

                    if missing_columns:

                        messages.error(
                            request,
                            "Missing required columns: "
                            + ", ".join(missing_columns)
                        )

                        workbook.close()
                        return redirect("..")

                    # -----------------------------------------
                    # Process rows in batches
                    # -----------------------------------------
                    batch = []

                    for row_number, values in enumerate(
                        rows_iterator,
                        start=2
                    ):

                        row = dict(
                            zip(headers, values)
                        )

                        batch.append(
                            (row_number, row)
                        )

                        if len(batch) >= BATCH_SIZE:

                            process_batch(batch)

                            batch.clear()

                    # -----------------------------------------
                    # Process remaining rows
                    # -----------------------------------------
                    if batch:
                        process_batch(batch)

                    workbook.close()

                except ValueError:

                    try:
                        workbook.close()
                    except Exception:
                        pass

                    return redirect("..")

                except Exception as e:

                    messages.error(
                        request,
                        f"Excel upload error: {str(e)}"
                    )

                    try:
                        workbook.close()
                    except Exception:
                        pass

                    return redirect("..")

            # =================================================
            # Invalid file
            # =================================================
            else:

                messages.error(
                    request,
                    "Upload only .csv or .xlsx files!"
                )

                return redirect("..")

            # =================================================
            # Final messages
            # =================================================
            messages.success(
                request,
                f"Upload complete! "
                f"Created: {created_count}, "
                f"Skipped: {skipped_count}"
            )

            return redirect("..")

        # =====================================================
        # GET
        # =====================================================
        form = UserBulkUploadForm()

        return render(
            request,
            "bulk_upload.html",
            {"form": form}
        )



# =====================================================================
# ✅ Ticket Admin
# =====================================================================

from django.contrib import admin
from .models import Ticket


@admin.register(Ticket)
class TicketAdmin(admin.ModelAdmin):

    # -------------------------------------------------------------
    # SHOW COLUMNS IN ADMIN TABLE
    # -------------------------------------------------------------
    list_display = (
        "ticket_number",
        "employee_code",
        "employee",
        "name",
        "location",
        "concern_category",
        "concern",
        "assigned_owner",
        "reassigned_to",
        "status",
        "created_at",
    )

    # -------------------------------------------------------------
    # FILTERS ON RIGHT SIDE
    # -------------------------------------------------------------
    list_filter = (
        "status",
        "location",
        "concern_category",
        "assigned_owner",
        "reassigned_to",
        "created_at",
    )

    # -------------------------------------------------------------
    # SEARCH BAR FIELDS
    # -------------------------------------------------------------
    search_fields = (
        "ticket_number",
        "employee_code",
        "name",
        "employee__username",
        "employee__employee_id",
        "location__name",
        "concern",
    )

    # -------------------------------------------------------------
    # READ ONLY FIELDS
    # (Auto-generated fields should NOT be editable)
    # -------------------------------------------------------------
    readonly_fields = (
        "ticket_number",
        "created_at",
        "updated_at",
    )

    # -------------------------------------------------------------
    # ORDERING
    # -------------------------------------------------------------
    ordering = ("-created_at",)




# =====================================================================
# ✅ Staff Performance Admin - with Bulk Upload
# =====================================================================

class BulkUploadForm(forms.Form):
    file = forms.FileField(label="Upload CSV/Excel File")


@admin.register(StaffPerformance)
class StaffPerformanceAdmin(admin.ModelAdmin):
    list_display = ("employee", "month", "rating", "earning_total", "deduction_total")
    search_fields = ("employee__username", "month")
    list_filter = ("month", "employee__role")
    ordering = ("-month",)
    change_list_template = "staff_performance_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("bulk-upload/", self.admin_site.admin_view(self.bulk_upload_view),
                 name="staffperformance_bulk_upload"),
        ]
        return custom_urls + urls

    def bulk_upload_view(self, request):
        if request.method == "POST":
            form = BulkUploadForm(request.POST, request.FILES)
            if form.is_valid():
                file = request.FILES["file"]

                try:
                    df = (
                        pd.read_csv(file)
                        if file.name.endswith(".csv")
                        else pd.read_excel(file)
                    )
                except Exception as e:
                    self.message_user(request, f"❌ Error reading file: {e}", level=messages.ERROR)
                    return redirect("..")

                required_columns = [
                    "employee_id", "month", "bau_status", "rating", "incentive",
                    "ot_sacoff_amount", "referral_bonus", "dsat_deduction",
                    "wrong_order_deduction", "mrd_deduction_staff",
                    "other_deduction", "earning_total", "deduction_total"
                ]

                missing = [c for c in required_columns if c not in df.columns]
                if missing:
                    self.message_user(request, f"❌ Missing columns: {', '.join(missing)}",
                                      level=messages.ERROR)
                    return redirect("..")

                created_count = 0

                for _, row in df.iterrows():
                    try:
                        user = CustomUser.objects.get(employee_id=row["employee_id"])
                        StaffPerformance.objects.update_or_create(
                            employee=user,
                            month=row["month"],
                            defaults={
                                "bau_status": row["bau_status"],
                                "rating": row["rating"],
                                "incentive": row["incentive"],
                                "ot_sacoff_amount": row["ot_sacoff_amount"],
                                "referral_bonus": row["referral_bonus"],
                                "dsat_deduction": row["dsat_deduction"],
                                "wrong_order_deduction": row["wrong_order_deduction"],
                                "mrd_deduction_staff": row["mrd_deduction_staff"],
                                "other_deduction": row["other_deduction"],
                                "earning_total": row["earning_total"],
                                "deduction_total": row["deduction_total"],
                            },
                        )
                        created_count += 1
                    except CustomUser.DoesNotExist:
                        continue

                self.message_user(request, f"✅ Uploaded {created_count} records successfully.",
                                  level=messages.SUCCESS)
                return redirect("..")

        else:
            form = BulkUploadForm()

        return render(request, "bulk_upload_form.html",
                      {"form": form, "title": "Bulk Upload Staff Performance", "opts": self.model._meta})


# =====================================================================
# ✅ Kitchen Log Admin
# =====================================================================

@admin.register(KitchenLog)
class KitchenLogAdmin(admin.ModelAdmin):
    list_display = ("emp_id", "emp_name", "location", "category", "created_at")
    search_fields = ("emp_id", "emp_name", "category")
    list_filter = ("location", "category", "created_at")
    ordering = ("-created_at",)


# =====================================================================
# ✅ Location Admin
# =====================================================================

@admin.register(Location)
class LocationAdmin(admin.ModelAdmin):
    list_display = ("code", "name")
    search_fields = ("code", "name")
    ordering = ("code",)


# =====================================================================
# ✅ Cluster Manager Profile Admin
# =====================================================================

class ClusterManagerProfileInline(admin.TabularInline):
    model = ClusterManagerProfile.locations.through
    extra = 1


@admin.register(ClusterManagerProfile)
class ClusterManagerProfileAdmin(admin.ModelAdmin):
    list_display = ("user",)
    inlines = [ClusterManagerProfileInline]
    filter_horizontal = ("locations",)
    search_fields = ("user__username",)


# =====================================================================
# ✅ Salary Slip Admin (Upload Excel/CSV)
# =====================================================================

class UploadFileForm(forms.Form):
    file = forms.FileField(label="Select Excel or CSV file")


@admin.register(SalarySlip)
class SalarySlipAdmin(admin.ModelAdmin):
    list_display = ('employee', 'month', 'year', 'net_pay', 'created_at')
    search_fields = ('employee__username', 'month', 'year')
    list_filter = ('year', 'month')
    change_list_template = "salaryslip_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('upload-salary/', self.admin_site.admin_view(self.upload_salary), name='upload_salary'),
        ]
        return custom + urls

    def upload_salary(self, request):
        if request.method == "POST":
            form = UploadFileForm(request.POST, request.FILES)
            if form.is_valid():
                file = form.cleaned_data['file']

                try:
                    df = pd.read_excel(file) if file.name.endswith(('.xls', '.xlsx')) else pd.read_csv(file)
                except Exception as e:
                    messages.error(request, f"❌ Error reading file: {e}")
                    return redirect("..")

                required_columns = [
                    'employee_id', 'month', 'year', 'present_days', 'lop_days',
                    'sac_off_ot', 'rating_incentive', 'km_mrd_incentive', 'arrears',
                    'referral_bonus', 'mrd_deduction', 'km_mrd_deduction',
                    'photo_deduction', 'missing_item_deduction', 'net_pay'
                ]

                missing_cols = [c for c in required_columns if c not in df.columns]
                if missing_cols:
                    messages.error(request, f"❌ Missing: {', '.join(missing_cols)}")
                    return redirect("..")

                created, skipped = 0, []

                for _, row in df.iterrows():
                    emp_id = str(row['employee_id']).strip()
                    user = User.objects.filter(employee_id=emp_id).first()

                    if not user:
                        skipped.append(emp_id)
                        continue

                    SalarySlip.objects.create(
                        employee=user,
                        month=row['month'],
                        year=row['year'],
                        present_days=row.get('present_days', 0) or 0,
                        lop_days=row.get('lop_days', 0) or 0,
                        sac_off_ot=row.get('sac_off_ot', 0) or 0,
                        rating_incentive=row.get('rating_incentive', 0) or 0,
                        km_mrd_incentive=row.get('km_mrd_incentive', 0) or 0,
                        arrears=row.get('arrears', 0) or 0,
                        referral_bonus=row.get('referral_bonus', 0) or 0,
                        mrd_deduction=row.get('mrd_deduction', 0) or 0,
                        km_mrd_deduction=row.get('km_mrd_deduction', 0) or 0,
                        photo_deduction=row.get('photo_deduction', 0) or 0,
                        missing_item_deduction=row.get('missing_item_deduction', 0) or 0,
                        net_pay=row.get('net_pay', 0) or 0,
                    )
                    created += 1

                msg = f"✅ {created} salary slips uploaded."
                if skipped:
                    msg += f" ⚠️ Skipped {len(skipped)} invalid employee IDs: {', '.join(skipped)}"

                messages.success(request, msg)
                return redirect("..")

        return render(request, 'upload_salary.html', {"form": UploadFileForm(), "title": "Upload Salary Slips"})


# =====================================================================
# ✅ Order Photo Admin
# =====================================================================

@admin.register(OrderPhoto)
class OrderPhotoAdmin(admin.ModelAdmin):
    list_display = ("order_id", "uploaded_by", "location", "uploaded_at")
    search_fields = ("order_id", "uploaded_by__username")
    list_filter = ("location", "uploaded_at")
    readonly_fields = ("image_preview",)

    def image_preview(self, obj):
        if obj.photo:
            return format_html(
                '<img src="{}" style="max-width:300px;border-radius:6px;">', obj.photo.url
            )
        return "No Image"

    image_preview.short_description = "Photo Preview"


